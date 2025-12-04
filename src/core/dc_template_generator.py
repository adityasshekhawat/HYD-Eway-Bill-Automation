#!/usr/bin/env python3
"""
DC Template Generator - Creates DCs in local Excel files with exact template format
"""

import pandas as pd
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins # For setting margins
from datetime import datetime
import time
import json # For state management
from decimal import Decimal, ROUND_HALF_UP # For accurate financial calculations
import decimal # For handling decimal operations
import re
try:
    from num2words import num2words # To convert numbers to words
except ImportError:
    import subprocess
    import sys
    # Use sys.executable to ensure we use the same python interpreter
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'num2words'])
    from num2words import num2words

from openpyxl.drawing.image import Image as ExcelImage

try:
    from PIL import Image as PILImage
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'Pillow'])
    from PIL import Image as PILImage

# Output directory for generated DCs
OUTPUT_DIR = "generated_dcs"
INPUT_DIR = "input_data"

# Import dynamic configuration loader
try:
    from .dynamic_hub_constants import get_dynamic_hub_constants, generate_hub_constants, generate_facility_mapping
except ImportError:
    # Fallback for standalone execution
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from core.dynamic_hub_constants import get_dynamic_hub_constants, generate_hub_constants, generate_facility_mapping

# Hub-specific constants with dynamic hub metadata integration
def extract_pincode_from_address(address):
    """Extract pincode from address string using regex"""
    pincode_match = re.search(r'(\d{6})', address)
    return pincode_match.group(1) if pincode_match else '000000'  # Changed from hardcoded 562123

# ✅ DYNAMIC CONFIGURATION: Generate hub constants from data files
# This replaces hardcoded Bangalore-specific values with data-driven configuration
HUB_CONSTANTS = generate_hub_constants()

# ✅ DYNAMIC CONFIGURATION: Generate facility mapping from data files  
# This replaces hardcoded facility addresses with data-driven configuration
FACILITY_ADDRESS_MAPPING = generate_facility_mapping()

# --- State Management ---
# REMOVED: Legacy sequence management - now using DCSequenceManager

def apply_formatting(ws, product_rows_count):
    """Apply visual formatting to the DC template using openpyxl."""
    
    # --- Define Styles ---
    title_font = Font(name='Calibri', size=14, bold=True)
    header_font = Font(name='Calibri', size=11, bold=True)
    footer_font = Font(name='Calibri', size=10, italic=True)
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    right_alignment = Alignment(horizontal='right', vertical='center')
    currency_format = '₹ #,##0.00'
    
    thin_border_side = Side(style='thin')
    box_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    product_header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # --- Column Widths ---
    column_widths = {'A': 15, 'B': 40, 'C': 15, 'D': 20, 'E': 15, 'F': 10, 'G': 12, 'H': 12, 'I': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # --- Row Heights (to better match the template) ---
    ws.row_dimensions[1].height = 30  # Title row
    ws.row_dimensions[2].height = 10  # Spacer row
    ws.row_dimensions[7].height = 25  # "Details from Dispatch" header
    ws.row_dimensions[8].height = 22  # Name row
    ws.row_dimensions[9].height = 45  # Address row (e.g., for 3 lines of text)
    ws.row_dimensions[10].height = 22 # GSTIN row
    ws.row_dimensions[11].height = 22 # FSSAI row
    ws.row_dimensions[12].height = 22 # State row
    ws.row_dimensions[13].height = 25 # Product table header

    # --- Merge Cells & Main Title ---
    ws.merge_cells('A1:G1')  # Changed from A1:I1 to make room for logo
    title_cell = ws['A1']
    title_cell.value = "Delivery Challan"
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    
    # Merge cells for company logo
    ws.merge_cells('H1:I1')
    
    # Apply border to the merged A1:G1 range manually for robustness
    thin_side = Side(style='thin')
    for col_letter in "ABCDEFG":  # Changed from "ABCDEFGHI" to "ABCDEFG"
        cell = ws[f'{col_letter}1']
        
        # Get existing border to not overwrite other sides
        current_border = cell.border
        left = current_border.left
        right = current_border.right
        
        if col_letter == 'A':
            left = thin_side
        if col_letter == 'G':  # Changed from 'I' to 'G'
            right = thin_side
            
        cell.border = Border(top=thin_side, bottom=thin_side, left=left, right=right)
    
    # Apply border to logo area H1:I1
    for col_letter in "HI":
        cell = ws[f'{col_letter}1']
        
        # Get existing border to not overwrite other sides
        current_border = cell.border
        left = current_border.left
        right = current_border.right
        
        if col_letter == 'H':
            left = thin_side
        if col_letter == 'I':
            right = thin_side
            
        cell.border = Border(top=thin_side, bottom=thin_side, left=left, right=right)
    
    # Add company logo
    add_company_logo(ws)

    # --- Static Headers Formatting ---
    ws.merge_cells('A7:D7')
    ws.merge_cells('E7:I7')
    headers_info = {
        'A3': "Serial no. of Challan:", 'D3': "Transport Mode:",
        'A4': "Date of Challan:", 'D4': "Vehicle number:",
        'A5': "Place of Supply :", 'A6': "State:",
        'A7': "Details from Dispatch", 'E7': "Details of Receiver",
        'A8': "Name:", 'E8': "Name:",
        'A9': "Address:", 'E9': "Address:",
        'A10': "GSTIN/UIN:", 'E10': "GSTIN/UIN:",
        'A11': "FSSAI:", 'E11': "FSSAI:",
        'A12': "State:", 'C12': "State Code:", 
        'E12': "State:", 'G12': "State Code:"
    }
    for cell_ref, text in headers_info.items():
        if ws[cell_ref].value:
            ws[cell_ref].font = header_font
            if cell_ref in ['A7', 'E7']:
                 ws[cell_ref].alignment = center_alignment

    # --- Sender & Receiver Details Formatting (Merging, Alignment, Borders) ---

    # Define the ranges for labels and data
    sender_labels = [f'A{i}' for i in range(8, 13)] + ['C12']
    receiver_labels = [f'E{i}' for i in range(8, 13)] + ['G12']
    
    sender_data_merges = {'B8': 'D8', 'B9': 'D9', 'B10': 'D10', 'B11': 'D11'}
    receiver_data_merges = {'F8': 'I8', 'F9': 'I9', 'F10': 'I10', 'F11': 'I11'}

    # Merge sender data cells
    for start, end in sender_data_merges.items():
        ws.merge_cells(f'{start}:{end}')
        
    # Merge receiver data cells
    for start, end in receiver_data_merges.items():
        ws.merge_cells(f'{start}:{end}')

    # Apply borders and alignment to the entire details section
    # Loop through rows 7-12 and columns A-I to apply borders
    for row in range(7, 13):
        for col_idx in range(1, 10): # A=1, I=9
            cell = ws.cell(row=row, column=col_idx)
            cell.border = box_border
            # Apply alignment to data cells (rows 8 and below)
            if row > 7:
                col_letter = get_column_letter(col_idx)
                # Data cells for sender (B, C, D) and receiver (F, G, H, I)
                if col_letter in ['B', 'C', 'D', 'F', 'G', 'H', 'I']:
                    cell.alignment = left_alignment

    # Re-apply bold font to all labels
    for cell_ref in sender_labels + receiver_labels:
        ws[cell_ref].font = header_font

    # --- Product Table Header ---
    product_table_start_row = 13
    product_header_cells = ws[f'A{product_table_start_row}:I{product_table_start_row}'][0]
    for cell in product_header_cells:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = product_header_fill
        cell.border = box_border
        
    # --- Product Table Data & Totals Borders ---
    totals_row_index = product_table_start_row + product_rows_count
    for row_idx in range(product_table_start_row + 1, totals_row_index + 1):
        ws.row_dimensions[row_idx].height = 20 # Increased height for product rows
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = box_border
            if col_idx == 1: # S.No.
                cell.alignment = right_alignment 
            elif col_idx == 2: # Product Description
                cell.alignment = left_alignment
            elif col_idx == 3: # HSN Code
                cell.alignment = left_alignment 
            elif col_idx in [4, 5, 6, 7, 8, 9]: # Numeric columns
                 cell.alignment = right_alignment
    ws[f'A{totals_row_index}'].font = bold_font

    # --- Totals Row Specific Formatting ---
    ws.merge_cells(f'A{totals_row_index}:C{totals_row_index}')
    total_label_cell = ws[f'A{totals_row_index}']
    total_label_cell.font = bold_font
    total_label_cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # Apply currency format to totals
    ws[f'E{totals_row_index}'].number_format = currency_format
    ws[f'G{totals_row_index}'].number_format = currency_format
    ws[f'H{totals_row_index}'].number_format = currency_format
    ws[f'I{totals_row_index}'].number_format = currency_format

    # --- Footer Section Formatting ---
    footer_start_row = totals_row_index + 1
    # Total in words
    ws.merge_cells(f'A{footer_start_row}:H{footer_start_row}')
    ws[f'A{footer_start_row}'].font = bold_font
    
    grand_total_cell = ws[f'I{footer_start_row}']
    grand_total_cell.font = bold_font
    grand_total_cell.alignment = right_alignment
    grand_total_cell.number_format = currency_format

    # Certification text
    footer_row_cert = footer_start_row + 1
    ws.merge_cells(f'A{footer_row_cert}:I{footer_row_cert}')
    ws[f'A{footer_row_cert}'].font = footer_font
    ws[f'A{footer_row_cert}'].alignment = left_alignment
    
    # Reasons for transportation
    footer_row_reasons = footer_start_row + 2
    ws.merge_cells(f'A{footer_row_reasons}:D{footer_row_reasons}')
    ws.merge_cells(f'E{footer_row_reasons}:I{footer_row_reasons}')
    ws[f'E{footer_row_reasons}'].font = bold_font
    ws[f'E{footer_row_reasons}'].alignment = center_alignment
    
    # Terms and Conditions
    footer_row_terms = footer_start_row + 4 # Row 86 relative to totals
    ws.merge_cells(f'A{footer_row_terms}:I{footer_row_terms+2}') # Assuming it spans 3 rows
    ws[f'A{footer_row_terms}'].alignment = left_alignment

    # Signatures
    footer_row_sig = footer_start_row + 7 # Row 89 relative to totals
    ws.merge_cells(f'E{footer_row_sig}:I{footer_row_sig}')
    ws[f'E{footer_row_sig}'].alignment = center_alignment

    # --- Page Setup for Printing ---
    # Calculate the last row of content
    last_content_row = totals_row_index + 8

    # Apply a thick outer border to the entire print area
    medium_side = Side(style='medium')
    for row in ws.iter_rows(min_row=1, max_row=last_content_row, min_col=1, max_col=9):
        for cell in row:
            # Start with a copy of the cell's existing border
            new_border = Border(left=cell.border.left,
                                  right=cell.border.right,
                                  top=cell.border.top,
                                  bottom=cell.border.bottom)
            # Apply the medium border to the edges of the print area
            if cell.row == 1:
                new_border.top = medium_side
            if cell.row == last_content_row:
                new_border.bottom = medium_side
            if cell.column == 1:
                new_border.left = medium_side
            if cell.column == 9:
                new_border.right = medium_side
            cell.border = new_border

    ws.print_area = f'A1:I{last_content_row}'  # Define the print area
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1  # Fit to one page wide
    ws.page_setup.fitToHeight = 0 # Allow as many pages tall as needed (0 means automatic)

    # Set print titles (repeat product header row on each page)
    # Product table header is on row 13
    ws.print_title_rows = '13:13' 

    # Set page margins (values are in inches)
    ws.page_margins = PageMargins(left=0.7, right=0.7, top=0.75, bottom=0.75, header=0.3, footer=0.3)


def create_dc_template(ws, product_data_start_row):
    """Create DC template structure in an openpyxl worksheet, including footer."""
    # Header section is populated in populate_dc_data to avoid writing labels twice.
    # We only set the product header here.
    product_header_values = ["S. No.", "Product Description", "HSN code", "Qty", "Taxable Value", 
                             "GST Rate", "CGST Amount", "SGST Amount", "Cess"]
    product_header_row = product_data_start_row
    for col_idx, value in enumerate(product_header_values, 1):
        ws.cell(row=product_header_row, column=col_idx, value=value)
    
    # --- Footer Section Labels ---
    # This assumes product data ends, then totals row, then this footer.
    # The exact row is calculated dynamically in populate_dc_data.
    return product_header_row + 1 # Return start row for product data


def populate_dc_data(ws, dc_data):
    """Populate all DC data, including header and footer, into an openpyxl worksheet."""
    hub_key = dc_data.get('hub_type') # e.g., 'AMOLAKCHAND'
    hub_details = HUB_CONSTANTS.get(hub_key, {})

    # ✅ CITY-AGNOSTIC: Use dynamic hub data if available, no hardcoded fallbacks
    place_of_supply = dc_data.get('place_of_supply', hub_details.get('place_of_supply', ''))
    hub_state = dc_data.get('hub_state', hub_details.get('state', ''))
    hub_state_code = dc_data.get('hub_state_code', hub_details.get('state_code', ''))
    
    print(f"🏢 DC Template using dynamic data:")
    print(f"   Place of Supply: {place_of_supply}")
    print(f"   Hub State: {hub_state} ({hub_state_code})")

    # CRITICAL FIX: Use dynamic facility address instead of hardcoded address
    facility_name = dc_data.get('facility_name', 'Unknown')
    if dc_data.get('facility_address'):
        # Use facility-specific address from DC data
        sender_address = dc_data['facility_address']
        print(f"🏢 DC Template: Using facility-specific address for {facility_name}")
        print(f"   Address: {sender_address}")
    else:
        # Fallback to hardcoded address
        sender_address = hub_details.get('sender_address', '')
        print(f"⚠️ DC Template: Using fallback address (facility address not available)")
        print(f"   Available DC fields: {list(dc_data.keys())}")
        print(f"   facility_address: {dc_data.get('facility_address', 'MISSING')}")

    # --- Populate Header ---
    headers_info = {
        'A3': "Serial no. of Challan:", 'C3': dc_data['serial_number'],
        'D3': "Transport Mode:", 'I3': "By Road",
        'A4': "Date of Challan:", 'C4': dc_data['date'].strftime('%d-%b-%Y'),
        'D4': "Vehicle number:",
        'A5': "Place of Supply :", 'C5': place_of_supply,  # Dynamic place of supply
        'A6': "State:", 'C6': hub_state,  # Dynamic state
        'A7': "Details from Dispatch", 'E7': "Details of Receiver",
        'A8': "Name:", 'B8': dc_data['sender_name'], 
        'E8': "Name:", 'F8': dc_data['receiver_name'],
        'A9': "Address:", 'B9': sender_address,  # FIXED: Use dynamic facility address
        'E9': "Address:", 'F9': dc_data.get('hub_address', ''), # Use hub_address from formatted data
        'A10': "GSTIN/UIN:", 'B10': hub_details.get('sender_gstin', ''),
        'E10': "GSTIN/UIN:", 'F10': hub_details.get('sender_gstin', ''),
        'A11': "FSSAI:", 'B11': '',
        'E11': "FSSAI:", 'F11': '',
        'A12': "State:", 'B12': hub_state, 'C12': "State Code:", 'D12': hub_state_code,  # Dynamic state info
        'E12': "State:", 'F12': hub_state,  # Dynamic state
        'G12': "State Code:", 'H12': hub_state_code  # Dynamic state code
    }
    for cell_ref, value in headers_info.items():
        ws[cell_ref] = value

    # --- Populate Products ---
    product_data_start_row = 14
    current_row = product_data_start_row
    total_qty = Decimal('0')
    total_taxable_value = Decimal('0')
    total_cgst = Decimal('0')
    total_sgst = Decimal('0')
    total_cess = Decimal('0')
    
    # Define a Decimal for rounding to two places
    TWO_PLACES = Decimal('0.01')

    for idx, product in enumerate(dc_data['products'], 1):
        value = product['Value'] # This is already a Decimal
        gst_rate = Decimal(str(product['GST Rate']))
        
        cgst_amount = (value * gst_rate * Decimal('0.5') / Decimal('100')).quantize(TWO_PLACES, rounding=ROUND_HALF_UP)
        sgst_amount = cgst_amount 
        
        # CESS CALCULATION FIX: The 'Cess' field contains CESS RATE, not amount
        # CESS Amount = (CESS Rate × Taxable Value) / 100
        cess_rate = product['Cess']  # This is the CESS rate (e.g., 12.0%)
        cess_amount = (cess_rate * value / Decimal('100')).quantize(TWO_PLACES, rounding=ROUND_HALF_UP)
        
        ws.cell(row=current_row, column=1, value=idx)
        ws.cell(row=current_row, column=2, value=product['Description'])
        ws.cell(row=current_row, column=3, value=product['HSN'])
        # Convert Decimals back to float for openpyxl if needed, or keep as string
        ws.cell(row=current_row, column=4, value=float(product['Quantity']))
        ws.cell(row=current_row, column=5, value=float(value))
        ws.cell(row=current_row, column=6, value=float(gst_rate))
        ws.cell(row=current_row, column=7, value=float(cgst_amount))
        ws.cell(row=current_row, column=8, value=float(sgst_amount))
        ws.cell(row=current_row, column=9, value=float(cess_amount))
        
        total_qty += Decimal(str(product['Quantity']))
        total_taxable_value += value
        total_cgst += cgst_amount
        total_sgst += sgst_amount
        total_cess += cess_amount
        current_row += 1
        
    # --- Populate Totals Row ---
    ws.cell(row=current_row, column=1, value="Total")
    ws.cell(row=current_row, column=4, value=float(total_qty))
    ws.cell(row=current_row, column=5, value=float(total_taxable_value))
    ws.cell(row=current_row, column=7, value=float(total_cgst))
    ws.cell(row=current_row, column=8, value=float(total_sgst))
    ws.cell(row=current_row, column=9, value=float(total_cess))
    
    product_rows_count = current_row - product_data_start_row
    
    # --- Footer Section ---
    footer_start_row = current_row + 1
    grand_total = (total_taxable_value + total_cgst + total_sgst + total_cess).quantize(TWO_PLACES, rounding=ROUND_HALF_UP)
    
    # Debug print to see the actual value
    print(f"DEBUG - grand_total: {grand_total} (type: {type(grand_total)})")
    
    # Convert grand_total to float and ensure it's a valid number before passing to num2words
    try:
        # First convert to string to handle Decimal objects properly
        grand_total_str = str(grand_total)
        grand_total_float = float(grand_total_str)
        amount_in_words = num2words(grand_total_float, to='currency', lang='en_IN', currency='INR').title()
    except (ValueError, decimal.InvalidOperation) as e:
        print(f"Warning: Could not convert amount to words: {e}")
        amount_in_words = "Amount conversion error"
    
    ws[f'A{footer_start_row}'] = f"Total in Words: {amount_in_words} Only"
    ws[f'I{footer_start_row}'] = float(grand_total)

    # Row 83 equivalent
    ws[f'A{footer_start_row+1}'] = "Certified that the particulars given above are true and correct"

    # Row 84 equivalent
    ws[f'A{footer_start_row+2}'] = "Reasons for transportation other than by way of supply: Intrastate Stock transfer between units of same entity"
    ws[f'E{footer_start_row+2}'] = hub_details['company_name']
    
    # Row 86 equivalent
    ws[f'A{footer_start_row+4}'] = "Terms & Conditions"
    
    # Row 89 equivalent
    ws[f'A{footer_start_row+7}'] = "Signature of Receiver"
    ws[f'E{footer_start_row+7}'] = "Authorised signatory"

    return product_rows_count + 1 # Include totals row for formatting

def create_dc_excel(dc_data_item):
    """Create a new DC Excel workbook, populate, format, and save it."""
    try:
        wb = Workbook()
        ws = wb.active
        
        # Use new DC sequence manager instead of legacy state
        try:
            from .dc_sequence_manager import dc_sequence_manager
            hub_key = dc_data_item.get('hub_type')
            facility_name = dc_data_item.get('facility_name', 'Arihant')  # Default to Arihant
            dc_number = dc_sequence_manager.generate_dc_number(hub_key, facility_name)
            print(f"✅ Generated new format DC number: {dc_number}")
        except Exception as e:
            print(f"❌ Error generating new format DC: {e}")
            dc_number = "LEGACY_DC_001"  # Fallback
        
        ws.title = f"DC_{dc_number}"

        # Set product header
        product_data_start_row = create_dc_template(ws, 13)
        
        # Populate all data (header, products, footer)
        dc_data_item['serial_number'] = dc_number
        product_rows_count = populate_dc_data(ws, dc_data_item)
        
        # Apply formatting
        apply_formatting(ws, product_rows_count)
        
        # Save the workbook
        if not os.path.exists(OUTPUT_DIR):
            os.makedirs(OUTPUT_DIR)
            
        # Include trip_ref_number in the filename for better identification
        trip_ref_number = dc_data_item.get('trip_ref_number', '').replace(',', '')  # Remove commas from trip_ref_number
        file_path = os.path.join(OUTPUT_DIR, f"{ws.title}_Trip{trip_ref_number}.xlsx")
        wb.save(file_path)
        
        # Clean up temporary logo files
        cleanup_temp_files(ws)
        
        print(f"✅ Created and Formatted DC: {dc_number} for Trip {trip_ref_number} at {file_path}")
        return True
        
    except Exception as e:
        print(f"❌ General Error creating DC Excel for {dc_number}: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def read_dc_data(input_file=None):
    """Read DC data from local CSV file"""
    try:
        from local_data_manager import get_dc_data
        
        # Default to Raw_DC.csv if no input file is provided
        if input_file is None:
            input_file = os.path.join("data", 'Raw_DC.csv')
            print(f"ℹ️ No input file provided. Defaulting to {input_file}")
            
        dc_data_list = get_dc_data(input_file)
        
        if dc_data_list is None:
            print("❌ Failed to get DC data")
            return None
            
        return dc_data_list
        
    except Exception as e:
        print(f"❌ Error reading DC data: {str(e)}")
        return None

def add_company_logo(ws, logo_path="/Users/jumbotail/Desktop/e-way bill /image.png"):
    """Add company logo to the Excel template in cells H1:I1"""
    try:
        if not os.path.exists(logo_path):
            print(f"⚠️ Logo file not found at {logo_path}")
            return False
        
        # Create a resized version of the logo in memory
        with PILImage.open(logo_path) as img:
            # Calculate ideal size for H1:I1 cells (approximately 120x30 pixels for Excel)
            # Excel column width is roughly 64 pixels for standard width
            # H1:I1 spans 2 columns, so about 128 pixels width
            # Row height is about 20 pixels for standard height
            target_width = 120
            target_height = 30
            
            # Calculate scaling to maintain aspect ratio
            original_ratio = img.width / img.height
            if target_width / target_height > original_ratio:
                # Height is the limiting factor
                new_height = target_height
                new_width = int(target_height * original_ratio)
            else:
                # Width is the limiting factor
                new_width = target_width
                new_height = int(target_width / original_ratio)
            
            # Resize image
            resized_img = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
            
            # Save to temporary file that will persist until Excel saves
            temp_logo_path = f"temp_logo_resized_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.png"
            resized_img.save(temp_logo_path, 'PNG')
        
        # Add image to Excel
        logo_img = ExcelImage(temp_logo_path)
        
        # Position the image in H1 (Excel coordinates start at 1,1)
        logo_img.anchor = 'H1'
        
        # Add to worksheet
        ws.add_image(logo_img)
        
        # Store temp file path for cleanup later
        if not hasattr(ws, '_temp_files'):
            ws._temp_files = []
        ws._temp_files.append(temp_logo_path)
        
        print("✅ Company logo added successfully")
        return True
        
    except Exception as e:
        print(f"❌ Error adding company logo: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def cleanup_temp_files(ws):
    """Clean up temporary files created during logo insertion"""
    try:
        if hasattr(ws, '_temp_files'):
            for temp_file in ws._temp_files:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                    print(f"🧹 Cleaned up temporary file: {temp_file}")
            ws._temp_files = []
    except Exception as e:
        print(f"⚠️ Warning: Could not clean up temporary files: {str(e)}")

def get_hub_pincode_from_address(hub_address):
    """Extract pincode from hub address"""
    if not hub_address:
        return '000000'  # ✅ CITY-AGNOSTIC: Changed from hardcoded 562123
    
    pincode_match = re.search(r'(\d{6})', hub_address)
    return pincode_match.group(1) if pincode_match else '000000'  # ✅ CITY-AGNOSTIC: Changed from hardcoded 562123

def main():
    """Main entry point"""
    try:
        print("🚀 Starting DC generation (Local Excel)...")
        
        # Read DC data
        input_file = sys.argv[1] if len(sys.argv) > 1 else None
        dc_data_list = read_dc_data(input_file)
        
        if not dc_data_list:
            print("❌ No DC data to process. Exiting.")
            return
            
        print(f"✅ Read data for {len(dc_data_list)} DCs")
        
        # Create output directory if it doesn't exist
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        
        # Generate DCs using new sequence manager
        for dc_data in dc_data_list:
            create_dc_excel(dc_data)  # No longer need to pass state
            
        print("✅ DC generation complete!")
        
    except Exception as e:
        print(f"❌ Error in main: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
