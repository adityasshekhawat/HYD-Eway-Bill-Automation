#!/usr/bin/env python3
"""
Vehicle DC Generator - Creates vehicle-based DCs using existing template format
Extends the existing DC creation logic for vehicle-based grouping
Enhanced with unified E-Way Bill template generation
"""

import os
import sys
from openpyxl import Workbook
from datetime import datetime
import json
from decimal import Decimal, ROUND_HALF_UP
try:
    from num2words import num2words
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'num2words'])
    from num2words import num2words

# Import existing formatting functions
from .dc_template_generator import (
    apply_formatting, create_dc_template, 
    HUB_CONSTANTS, add_company_logo, cleanup_temp_files
)

# Import e-way bill generation components
try:
    from eway_bill.eway_bill_template_generator import EwayBillTemplateGenerator
except ImportError:
    # Fallback for when running from different contexts or when e-way module is not available
    try:
        import sys
        import os
        sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
        from eway_bill.eway_bill_template_generator import EwayBillTemplateGenerator
    except ImportError:
        # If e-way bill module is not available, create a dummy class
        class EwayBillTemplateGenerator:
            def generate_template_from_dc(self, dc_data):
                return []
            def save_to_excel(self, rows, file_path):
                pass

# Import PDF generation components
try:
    from pdf_generator.dc_pdf_generator import create_dc_pdf_from_excel_data
except ImportError:
    # Fallback for when running from different contexts
    import sys
    import os
    sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
    from pdf_generator.dc_pdf_generator import create_dc_pdf_from_excel_data

# Output directory for vehicle-based DCs
VEHICLE_OUTPUT_DIR = "generated_vehicle_dcs"

def populate_vehicle_dc_data(ws, dc_data):
    """Populate DC data with vehicle number in cell I4 and all tax information"""
    hub_key = dc_data.get('hub_type') # e.g., 'AMOLAKCHAND'
    hub_details = HUB_CONSTANTS.get(hub_key, {})

    # --- Populate Header ---
    headers_info = {
        'A3': "Serial no. of Challan:", 'C3': dc_data['serial_number'],
        'D3': "Transport Mode:", 'I3': "By Road",
        'A4': "Date of Challan:", 'C4': dc_data['date'].strftime('%d-%b-%Y'),
        'D4': "Vehicle number:", 'I4': dc_data.get('vehicle_number', ''),  # VEHICLE NUMBER HERE!
        'A5': "Place of Supply :", 'C5': hub_details.get('place_of_supply', ''),
        'A6': "State:", 'C6': hub_details.get('state', ''),
        'A7': "Details from Dispatch", 'E7': "Details of Receiver",
        'A8': "Name:", 'B8': dc_data['sender_name'], 
        'E8': "Name:", 'F8': dc_data['receiver_name'],
        'A9': "Address:", 'B9': hub_details.get('sender_address', ''),
        'E9': "Address:", 'F9': dc_data.get('hub_address', ''), # Use hub_address from formatted data
        'A10': "GSTIN/UIN:", 'B10': hub_details.get('sender_gstin', ''),
        'E10': "GSTIN/UIN:", 'F10': hub_details.get('sender_gstin', ''),
        'A11': "FSSAI:", 'B11': '',
        'E11': "FSSAI:", 'F11': '',
        'A12': "State:", 'B12': hub_details.get('state', ''), 'C12': "State Code:", 'D12': hub_details.get('state_code', ''),
        'E12': "State:", 'F12': hub_details.get('state', ''),
        'G12': "State Code:", 'H12': hub_details.get('state_code', '')
    }
    for cell_ref, value in headers_info.items():
        ws[cell_ref] = value

    # --- Populate Products ---
    product_data_start_row = 14
    current_row = product_data_start_row
    total_qty = Decimal('0')
    total_taxable_value = Decimal('0')
    total_cgst = Decimal('0')
    total_sgst = Decimal('0')
    total_cess = Decimal('0')
    
    # Define a Decimal for rounding to two places
    TWO_PLACES = Decimal('0.01')

    for idx, product in enumerate(dc_data['products'], 1):
        value = Decimal(str(product['Value']))  # Ensure it's a Decimal
        gst_rate = Decimal(str(product.get('GST Rate', 0)))
        
        cgst_amount = (value * gst_rate * Decimal('0.5') / Decimal('100')).quantize(TWO_PLACES, rounding=ROUND_HALF_UP)
        sgst_amount = cgst_amount 
        
        # CESS CALCULATION FIX: The 'Cess' field contains CESS RATE, not amount
        # CESS Amount = (CESS Rate × Taxable Value) / 100
        cess_rate = Decimal(str(product.get('Cess', 0)))  # FIXED: Use get() with default 0 instead of direct access
        cess_amount = (cess_rate * value / Decimal('100')).quantize(TWO_PLACES, rounding=ROUND_HALF_UP)
        
        # Debug print to verify data and CESS calculation
        print(f"DEBUG Product {idx}: HSN={product.get('HSN', 'MISSING')}, GST={gst_rate}%, Value={value}")
        if cess_rate > 0:
            print(f"   CESS Rate: {cess_rate}%, CESS Amount: ₹{cess_amount}")
        
        # Get HSN code - ensure it's a string and not empty
        hsn_code = str(product.get('HSN', 'N/A'))
        if hsn_code == 'N/A' or not hsn_code.strip():
            # Try to get HSN from jpin if available
            if 'jpin' in product:
                print(f"   ⚠️ Missing HSN for product {idx}, trying to find by JPIN: {product['jpin']}")
                # This is just a placeholder - in a real implementation, you'd look up the HSN by JPIN
                hsn_code = 'N/A'
        
        # Set cell values with proper formatting
        ws.cell(row=current_row, column=1, value=idx)
        ws.cell(row=current_row, column=2, value=product['Description'])
        ws.cell(row=current_row, column=3, value=hsn_code)  # HSN code
        ws.cell(row=current_row, column=4, value=float(product['Quantity']))
        ws.cell(row=current_row, column=5, value=float(value))
        ws.cell(row=current_row, column=6, value=float(gst_rate))  # GST rate
        ws.cell(row=current_row, column=7, value=float(cgst_amount))
        ws.cell(row=current_row, column=8, value=float(sgst_amount))
        ws.cell(row=current_row, column=9, value=float(cess_amount))
        
        # Ensure HSN and GST rate are visible by setting number format
        ws.cell(row=current_row, column=3).number_format = '@'  # Text format for HSN
        ws.cell(row=current_row, column=6).number_format = '0.0'  # Show decimal for GST rate
        
        total_qty += Decimal(str(product['Quantity']))
        total_taxable_value += value
        total_cgst += cgst_amount
        total_sgst += sgst_amount
        total_cess += cess_amount
        current_row += 1
        
    # --- Populate Totals Row ---
    ws.cell(row=current_row, column=1, value="Total")
    ws.cell(row=current_row, column=4, value=float(total_qty))
    ws.cell(row=current_row, column=5, value=float(total_taxable_value))
    ws.cell(row=current_row, column=7, value=float(total_cgst))
    ws.cell(row=current_row, column=8, value=float(total_sgst))
    ws.cell(row=current_row, column=9, value=float(total_cess))
    
    product_rows_count = current_row - product_data_start_row
    
    # --- Footer Section ---
    footer_start_row = current_row + 1
    grand_total = (total_taxable_value + total_cgst + total_sgst + total_cess).quantize(TWO_PLACES, rounding=ROUND_HALF_UP)
    
    # Debug print to see the actual value
    print(f"DEBUG - grand_total: {grand_total} (type: {type(grand_total)})")
    
    # Convert grand_total to float and ensure it's a valid number before passing to num2words
    try:
        # First convert to string to handle Decimal objects properly
        grand_total_str = str(grand_total)
        grand_total_float = float(grand_total_str)
        amount_in_words = num2words(grand_total_float, to='currency', lang='en_IN', currency='INR').title()
    except (ValueError, decimal.InvalidOperation) as e:
        print(f"Warning: Could not convert amount to words: {e}")
        amount_in_words = "Amount conversion error"
    
    ws[f'A{footer_start_row}'] = f"Total in Words: {amount_in_words} Only"
    ws[f'I{footer_start_row}'] = float(grand_total)

    # Row 83 equivalent
    ws[f'A{footer_start_row+1}'] = "Certified that the particulars given above are true and correct"

    # Row 84 equivalent
    ws[f'A{footer_start_row+2}'] = "Reasons for transportation other than by way of supply: Intrastate Stock transfer between units of same entity"
    ws[f'E{footer_start_row+2}'] = hub_details['company_name']
    
    # Row 86 equivalent
    ws[f'A{footer_start_row+4}'] = "Terms & Conditions"
    
    # Row 89 equivalent
    ws[f'A{footer_start_row+7}'] = "Signature of Receiver"
    ws[f'E{footer_start_row+7}'] = "Authorised signatory"

    return product_rows_count + 1 # Include totals row for formatting

class VehicleDCGenerator:
    def __init__(self):
        # Initialize new sequence manager
        try:
            from .dc_sequence_manager import dc_sequence_manager
            self.new_sequence_manager = dc_sequence_manager
            print("✅ Initialized with new DC sequence manager")
        except ImportError as e:
            print(f"❌ Could not import new sequence manager: {e}")
            raise ImportError("DC sequence manager is required for operation")
    
    def generate_vehicle_dc_number(self, hub_type, vehicle_number, sequence_num=1, facility_name=None, hub_value=None):
        """
        Generate vehicle-based DC number using new naming convention
        
        Args:
            hub_type: Hub type (SOURCINGBEE, AMOLAKCHAND, BODEGA)
            vehicle_number: Vehicle number
            sequence_num: Sequence number for multiple DCs
            facility_name: Facility name for new format
            hub_value: Optional hub value (e.g., 'HYD_NCH') for hub-specific sequences
            
        Returns:
            DC number in new format (e.g., AKDCAH000001, AKDCHYDNCH00000001)
        """
        if not self.new_sequence_manager:
            raise RuntimeError("DC sequence manager not available")
            
        if not facility_name:
            facility_name = 'Arihant'  # Default facility
            
        dc_number = self.new_sequence_manager.generate_dc_number(hub_type, facility_name, hub_value)
        
        # Handle multiple DCs from same data
        if sequence_num > 1:
            # For multiple DCs, append sequence number to maintain uniqueness
            dc_number = f"{dc_number}_{sequence_num:02d}"
        
        print(f"🔢 Generated new format vehicle DC: {dc_number} (hub: {hub_value or 'N/A'})")
        return dc_number
    
    def create_vehicle_dc_excel(self, dc_data_item, output_dir=None, generate_eway_template=True, generate_pdf=True):
        """
        Create a vehicle-based DC Excel file with optional PDF and e-way bill template generation
        
        Args:
            dc_data_item: DC data dictionary
            output_dir: Output directory for files
            generate_eway_template: Whether to generate e-way bill template (default: True)
            generate_pdf: Whether to generate PDF version (default: True)
            
        Returns:
            Dictionary with DC, PDF, and e-way template file information
        """
        try:
            if output_dir is None:
                output_dir = VEHICLE_OUTPUT_DIR
                
            # Ensure output directory exists
            os.makedirs(output_dir, exist_ok=True)
            
            # ✅ NEW: Check company availability in state before generating DC
            hub_type = dc_data_item.get('hub_type', 'SOURCINGBEE')
            state = dc_data_item.get('hub_state') or dc_data_item.get('state', '')
            
            # Import config loader for availability check
            from .config_loader import get_config_loader
            config_loader = get_config_loader()
            
            if state and not config_loader.is_company_available_in_state(hub_type, state):
                print(f"⚠️ Skipping DC generation: {hub_type} not available in {state}")
                return {
                    'dc_number': None,
                    'file_path': None,
                    'vehicle_number': dc_data_item.get('vehicle_number', 'UNKNOWN'),
                    'trip_refs': dc_data_item.get('trip_refs', []),
                    'hub_type': hub_type,
                    'state': state,
                    'skipped': True,
                    'skip_reason': f'{hub_type} not available in {state}',
                    'pdf_generated': False,
                    'eway_template_generated': False
                }
            
            wb = Workbook()
            ws = wb.active
            
            # Generate DC number with enhanced facility detection
            vehicle_number = dc_data_item.get('vehicle_number', 'UNKNOWN')
            dc_sequence = dc_data_item.get('dc_sequence', 1)
            
            # NEW: Extract facility name from DC data
            facility_name = dc_data_item.get('facility_name', 'Arihant')  # Default to Arihant
            print(f"🏭 Using facility for DC generation: {facility_name}")
            
            # NEW: Extract hub information for Telangana hub-specific sequences
            hub_value = dc_data_item.get('hub', '')  # e.g., 'HYD_NCH', 'HYD_BAL'
            print(f"📍 Hub value for DC sequencing: {hub_value or 'N/A'}")
            
            # FIXED: Reserve DC number first (doesn't increment sequence yet)
            dc_number = self.new_sequence_manager.reserve_dc_number(hub_type, facility_name, hub_value)
            ws.title = f"DC_{dc_number}"
            
            # Prepare DC data in the format expected by existing functions
            formatted_dc_data = self._format_dc_data_for_template(dc_data_item, dc_number, vehicle_number)
            
            # Debug: Print first few products to verify data
            products = formatted_dc_data.get('products', [])
            print(f"\n🔍 DEBUG - Vehicle {vehicle_number} DC {dc_number}:")
            print(f"   Company: {hub_type}")
            print(f"   Facility: {facility_name}")
            print(f"   Products count: {len(products)}")
            for i, product in enumerate(products[:3]):
                print(f"   Product {i+1}: {product.get('Description', 'N/A')[:30]}...")
                print(f"     HSN: {product.get('HSN', 'MISSING')}")
                print(f"     GST Rate: {product.get('GST Rate', 'MISSING')}%")
                print(f"     Value: ₹{product.get('Value', 'MISSING')}")
            
            # Use existing template creation logic
            product_data_start_row = create_dc_template(ws, 13)
            
            # Use our custom populate function that includes vehicle number
            product_rows_count = populate_vehicle_dc_data(ws, formatted_dc_data)
            
            # Apply formatting
            apply_formatting(ws, product_rows_count)
            
            # Generate filename with vehicle number and new DC format
            trip_refs_str = "_".join(str(tid) for tid in dc_data_item.get('trip_refs', [])[:3])  # Limit length
            if len(trip_refs_str) > 50:  # Prevent overly long filenames
                trip_refs_str = f"{len(dc_data_item.get('trip_refs', []))}trips"
                
            filename = f"{ws.title}_Vehicle{vehicle_number}_{trip_refs_str}.xlsx"
            file_path = os.path.join(output_dir, filename)
            
            wb.save(file_path)
            
            # FIXED: Confirm DC number only after successful save
            if self.new_sequence_manager.confirm_dc_number(dc_number):
                print(f"✅ DC sequence confirmed for {dc_number}")
            else:
                print(f"⚠️ Failed to confirm DC sequence for {dc_number}")
            
            # Clean up temporary logo files
            cleanup_temp_files(ws)
            
            print(f"✅ Created Vehicle DC: {dc_number} for Vehicle {vehicle_number} at {file_path}")
            
            # Initialize result structure
            result = {
                'dc_number': dc_number,
                'file_path': file_path,
                'vehicle_number': vehicle_number,
                'trip_refs': dc_data_item.get('trip_refs', []),
                'hub_type': hub_type,
                'facility_name': facility_name,  # NEW: Include facility name in result
                'product_count': len(dc_data_item.get('products', [])),
                'dc_sequence': dc_sequence,
                'total_dcs': dc_data_item.get('total_dcs', 1),
                'pdf_generated': False,
                'pdf_path': None,
                'pdf_status': 'not_generated',
                'pdf_error': None,
                'eway_template_generated': False,
                'eway_template_path': None,
                'eway_template_status': 'not_generated',
                'eway_error': None
            }
            
            # Generate E-Way Bill Template if requested
            if generate_eway_template:
                try:
                    eway_result = self._generate_eway_template(formatted_dc_data, output_dir, dc_number, vehicle_number, trip_refs_str)
                    if eway_result:
                        result.update({
                            'eway_template_generated': True,
                            'eway_template_path': eway_result['file_path'],
                            'eway_template_status': 'success',
                            'eway_row_count': eway_result.get('row_count', 0)
                        })
                        print(f"✅ Generated E-Way template for DC {dc_number}: {eway_result['file_path']}")
                    else:
                        result.update({
                            'eway_template_status': 'failed',
                            'eway_error': 'E-Way template generation returned no result'
                        })
                        print(f"⚠️ E-Way template generation failed for DC {dc_number}")
                        
                except Exception as e:
                    error_msg = str(e)
                    result.update({
                        'eway_template_status': 'failed',
                        'eway_error': error_msg
                    })
                    print(f"⚠️ E-Way template generation failed for DC {dc_number}: {error_msg}")
                    # Don't fail the entire process - DC creation was successful
            
            # Generate PDF version if requested
            if generate_pdf:
                try:
                    # Generate PDF filename
                    pdf_filename = filename.replace('.xlsx', '.pdf')
                    pdf_file_path = os.path.join(output_dir, pdf_filename)
                    
                    # Create PDF using the same data as Excel
                    pdf_success = create_dc_pdf_from_excel_data(formatted_dc_data, pdf_file_path)
                    
                    if pdf_success:
                        result.update({
                            'pdf_generated': True,
                            'pdf_path': pdf_file_path,
                            'pdf_status': 'success'
                        })
                        print(f"✅ Generated PDF DC: {dc_number} for Vehicle {vehicle_number} at {pdf_file_path}")
                    else:
                        result.update({
                            'pdf_status': 'failed',
                            'pdf_error': 'PDF generation returned False'
                        })
                        print(f"⚠️ PDF generation failed for DC {dc_number}")
                        
                except Exception as e:
                    error_msg = str(e)
                    result.update({
                        'pdf_status': 'failed',
                        'pdf_error': error_msg
                    })
                    print(f"⚠️ PDF generation failed for DC {dc_number}: {error_msg}")
                    # Don't fail the entire process - Excel creation was successful
            
            return result
            
        except Exception as e:
            print(f"❌ Error creating vehicle DC Excel: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def _format_dc_data_for_template(self, dc_data_item, dc_number, vehicle_number):
        """Format vehicle DC data for existing template functions"""
        return {
            'serial_number': dc_number,
            'date': dc_data_item.get('date', datetime.now()),
            'hub_type': dc_data_item.get('hub_type', 'SOURCINGBEE'),
            'sender_name': dc_data_item.get('sender_name', ''),
            'receiver_name': dc_data_item.get('receiver_name', ''),
            'hub_address': dc_data_item.get('hub_address', ''),
            'hub_pincode': dc_data_item.get('hub_pincode', ''),  # Add hub pincode
            'hub_state': dc_data_item.get('hub_state', 'Karnataka'),  # Add hub state
            'hub_state_code': dc_data_item.get('hub_state_code', '29'),  # Add hub state code
            'distance': dc_data_item.get('distance', ''),  # FIXED: Changed from '100' to '' for blank delivered km
            'place_of_supply': dc_data_item.get('place_of_supply', ''),  # ✅ No fallback
            'hub_name': dc_data_item.get('hub_name', ''),  # Add hub name
            'facility_name': dc_data_item.get('facility_name', ''),  # Add facility name
            'facility_address': dc_data_item.get('facility_address', ''),  # CRITICAL FIX: Add facility address
            'facility_address_line1': dc_data_item.get('facility_address_line1', ''),  # Add facility address
            'facility_address_line2': dc_data_item.get('facility_address_line2', ''),
            'facility_pincode': dc_data_item.get('facility_pincode', ''),
            'facility_city': dc_data_item.get('facility_city', ''),
            'facility_state': dc_data_item.get('facility_state', ''),
            'products': dc_data_item.get('products', []),
            'vehicle_number': vehicle_number,
            'trip_refs': dc_data_item.get('trip_refs', [])
        }
    
    def _generate_eway_template(self, formatted_dc_data, output_dir, dc_number, vehicle_number, trip_refs_str):
        """
        Generate e-way bill template from DC data
        
        Args:
            formatted_dc_data: Formatted DC data dictionary
            output_dir: Output directory
            dc_number: DC number
            vehicle_number: Vehicle number
            trip_refs_str: Trip references string for filename
            
        Returns:
            Dictionary with e-way template generation result
        """
        try:
            # Initialize e-way bill template generator
            eway_generator = EwayBillTemplateGenerator()
            
            # Convert DC data to e-way format (formatted_dc_data is already in the right format)
            eway_dc_data = self._convert_dc_to_eway_format(formatted_dc_data)
            
            # Generate e-way bill template rows
            eway_rows = eway_generator.generate_template_from_dc(eway_dc_data)
            
            if not eway_rows:
                return None
            
            # Generate e-way template filename
            eway_filename = f"EWAY_{dc_number}_Vehicle{vehicle_number}_{trip_refs_str}.xlsx"
            eway_file_path = os.path.join(output_dir, eway_filename)
            
            # Save e-way template to Excel
            eway_generator.save_to_excel(eway_rows, eway_file_path)
            
            return {
                'file_path': eway_file_path,
                'row_count': len(eway_rows),
                'filename': eway_filename
            }
            
        except Exception as e:
            print(f"❌ Error generating e-way template: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def _convert_dc_to_eway_format(self, formatted_dc_data):
        """
        Convert formatted DC data to e-way bill format
        The formatted_dc_data from _format_dc_data_for_template is already mostly compatible
        """
        # The formatted_dc_data structure is already compatible with e-way generation
        # Just need to ensure the correct field mapping and pass all hub data
        return {
            'dc_number': formatted_dc_data.get('serial_number', ''),
            'date': formatted_dc_data.get('date', datetime.now()),
            'vehicle_number': formatted_dc_data.get('vehicle_number', ''),
            'sender_name': formatted_dc_data.get('sender_name', ''),
            'receiver_name': formatted_dc_data.get('receiver_name', ''),
            'hub_type': formatted_dc_data.get('hub_type', 'SOURCINGBEE'),
            'hub_address': formatted_dc_data.get('hub_address', ''),
            'hub_pincode': formatted_dc_data.get('hub_pincode', ''),  # Pass hub pincode
            'hub_state': formatted_dc_data.get('hub_state', 'Karnataka'),  # Pass hub state
            'hub_state_code': formatted_dc_data.get('hub_state_code', '29'),  # Pass hub state code
            'distance': formatted_dc_data.get('distance', ''),  # Pass distance
            'place_of_supply': formatted_dc_data.get('place_of_supply', ''),  # ✅ No fallback
            'hub_name': formatted_dc_data.get('hub_name', ''),  # Pass hub name
            'facility_name': formatted_dc_data.get('facility_name', ''),  # Pass facility name
            'facility_address_line1': formatted_dc_data.get('facility_address_line1', ''),  # Pass facility address
            'facility_address_line2': formatted_dc_data.get('facility_address_line2', ''),
            'facility_pincode': formatted_dc_data.get('facility_pincode', ''),
            'facility_city': formatted_dc_data.get('facility_city', ''),
            'facility_state': formatted_dc_data.get('facility_state', ''),
            'products': formatted_dc_data.get('products', []),
            'trip_refs': formatted_dc_data.get('trip_refs', [])
        }
    
    def _generate_consolidated_eway_templates(self, eway_consolidation_data, output_dir):
        """
        Generate consolidated e-way bill templates grouped by seller
        
        Args:
            eway_consolidation_data: Dictionary with seller -> list of DC data
            output_dir: Output directory for consolidated files
            
        Returns:
            Dictionary with seller -> consolidation result
        """
        if not output_dir:
            output_dir = VEHICLE_OUTPUT_DIR
        
        os.makedirs(output_dir, exist_ok=True)
        
        consolidation_results = {}
        
        try:
            # Initialize e-way bill template generator
            eway_generator = EwayBillTemplateGenerator()
            
            for seller, dc_entries in eway_consolidation_data.items():
                print(f"\n🔄 Consolidating e-way bills for {seller} ({len(dc_entries)} DCs)...")
                
                # Collect all e-way bill rows for this seller
                all_eway_rows = []
                dc_numbers = []
                vehicle_numbers = set()
                
                for dc_entry in dc_entries:
                    try:
                        # Format DC data for e-way bill generation
                        formatted_dc_data = self._format_dc_data_for_template(
                            dc_entry['dc_data'], 
                            dc_entry['dc_number'], 
                            dc_entry['vehicle_number']
                        )
                        
                        # Convert to e-way format
                        eway_dc_data = self._convert_dc_to_eway_format(formatted_dc_data)
                        
                        # Generate e-way bill template rows for this DC
                        eway_rows = eway_generator.generate_template_from_dc(eway_dc_data)
                        
                        if eway_rows:
                            all_eway_rows.extend(eway_rows)
                            dc_numbers.append(dc_entry['dc_number'])
                            vehicle_numbers.add(dc_entry['vehicle_number'])
                            print(f"   ✅ {dc_entry['dc_number']}: {len(eway_rows)} rows")
                        else:
                            print(f"   ⚠️ {dc_entry['dc_number']}: No rows generated")
                            
                    except Exception as e:
                        print(f"   ❌ Error processing DC {dc_entry['dc_number']}: {str(e)}")
                        continue
                
                if all_eway_rows:
                    # Generate consolidated filename - FIXED: Remove vehicle-specific naming for true global consolidation
                    consolidated_filename = f"EWAY_{seller}_Consolidated.xlsx"
                    consolidated_file_path = os.path.join(output_dir, consolidated_filename)
                    
                    # Save consolidated e-way template
                    try:
                        eway_generator.save_to_excel(all_eway_rows, consolidated_file_path)
                        
                        consolidation_results[seller] = {
                            'success': True,
                            'file_path': consolidated_file_path,
                            'total_rows': len(all_eway_rows),
                            'dc_count': len(dc_entries),
                            'dc_numbers': dc_numbers,
                            'vehicle_numbers': list(vehicle_numbers)
                        }
                        
                        print(f"   ✅ Consolidated {len(dc_entries)} DCs from {len(vehicle_numbers)} vehicles into {consolidated_filename}")
                        print(f"   📊 Total rows: {len(all_eway_rows)}")
                        
                    except Exception as e:
                        error_msg = f"Failed to save consolidated file: {str(e)}"
                        consolidation_results[seller] = {
                            'success': False,
                            'error': error_msg,
                            'dc_count': len(dc_entries),
                            'dc_numbers': dc_numbers
                        }
                        print(f"   ❌ {error_msg}")
                
                else:
                    error_msg = f"No e-way bill rows generated for {seller}"
                    consolidation_results[seller] = {
                        'success': False,
                        'error': error_msg,
                        'dc_count': len(dc_entries),
                        'dc_numbers': dc_numbers
                    }
                    print(f"   ❌ {error_msg}")
            
            return consolidation_results
            
        except Exception as e:
            print(f"❌ Error in consolidated e-way template generation: {str(e)}")
            import traceback
            traceback.print_exc()
            return {}
    
    def generate_vehicle_dcs(self, vehicle_dc_data_list, output_dir=None, generate_eway_templates=True, generate_pdfs=True, consolidate_eway_bills=False):
        """
        Generate multiple vehicle DCs with optional PDF and e-way template generation
        
        Args:
            vehicle_dc_data_list: List of vehicle DC data
            output_dir: Output directory for files
            generate_eway_templates: Whether to generate e-way templates (default: True)
            generate_pdfs: Whether to generate PDF versions (default: True)
            consolidate_eway_bills: Whether to consolidate e-way bills by seller (default: False)
            
        Returns:
            List of result dictionaries with DC, PDF, and e-way template information
        """
        if not vehicle_dc_data_list:
            print("❌ No vehicle DC data provided")
            return []
            
        # DEBUG: Add comprehensive logging for consolidation troubleshooting
        print(f"\n🔍 DEBUG - Consolidation Data Collection:")
        print(f"   Total input data entries: {len(vehicle_dc_data_list)}")
        print(f"   Consolidate e-way bills: {consolidate_eway_bills}")
        
        # Analyze input data structure
        unique_vehicles = set()
        for i, dc_data_entry in enumerate(vehicle_dc_data_list):
            if isinstance(dc_data_entry, list):
                print(f"   Entry {i}: LIST with {len(dc_data_entry)} DCs")
                for j, dc in enumerate(dc_data_entry):
                    vehicle_num = dc.get('vehicle_number', 'UNKNOWN')
                    hub_type = dc.get('hub_type', 'UNKNOWN')
                    unique_vehicles.add(vehicle_num)
                    print(f"     DC {j}: Vehicle {vehicle_num}, Hub {hub_type}")
            else:
                vehicle_num = dc_data_entry.get('vehicle_number', 'UNKNOWN')
                hub_type = dc_data_entry.get('hub_type', 'UNKNOWN')
                unique_vehicles.add(vehicle_num)
                print(f"   Entry {i}: SINGLE DC - Vehicle {vehicle_num}, Hub {hub_type}")
        
        print(f"   Unique vehicles in input: {list(unique_vehicles)}")
            
        results = []
        eway_success_count = 0
        eway_fail_count = 0
        pdf_success_count = 0
        pdf_fail_count = 0
        
        # NEW: E-way bill consolidation data collection
        eway_consolidation_data = {}
        
        try:
            for dc_data_list in vehicle_dc_data_list:
                # dc_data_list contains multiple DCs for one seller (if split due to 250 limit)
                if isinstance(dc_data_list, list):
                    for dc_data in dc_data_list:
                        # Determine whether to generate individual e-way templates
                        generate_individual_eway = generate_eway_templates and not consolidate_eway_bills
                        
                        result = self.create_vehicle_dc_excel(dc_data, output_dir, generate_individual_eway, generate_pdfs)
                        if result:
                            results.append(result)
                            
                            # NEW: Collect e-way bill data for consolidation
                            if generate_eway_templates and consolidate_eway_bills:
                                seller = dc_data.get('hub_type', 'UNKNOWN')
                                if seller not in eway_consolidation_data:
                                    eway_consolidation_data[seller] = []
                                eway_consolidation_data[seller].append({
                                    'dc_data': dc_data,
                                    'dc_number': result['dc_number'],
                                    'vehicle_number': result['vehicle_number'],
                                    'trip_refs': result.get('trip_refs', [])
                                })
                                print(f"   🔄 Added to consolidation: {seller} - DC {result['dc_number']} (Vehicle {result['vehicle_number']})")
                                
                            # Track individual e-way template generation success (if not consolidating)
                            if generate_individual_eway:
                                if result.get('eway_template_status') == 'success':
                                    eway_success_count += 1
                                elif result.get('eway_template_status') == 'failed':
                                    eway_fail_count += 1
                            # Track PDF generation success
                            if generate_pdfs:
                                if result.get('pdf_status') == 'success':
                                    pdf_success_count += 1
                                elif result.get('pdf_status') == 'failed':
                                    pdf_fail_count += 1
                else:
                    # Single DC
                    # Determine whether to generate individual e-way templates
                    generate_individual_eway = generate_eway_templates and not consolidate_eway_bills
                    
                    result = self.create_vehicle_dc_excel(dc_data_list, output_dir, generate_individual_eway, generate_pdfs)
                    if result:
                        results.append(result)
                        
                        # NEW: Collect e-way bill data for consolidation
                        if generate_eway_templates and consolidate_eway_bills:
                            seller = dc_data_list.get('hub_type', 'UNKNOWN')
                            if seller not in eway_consolidation_data:
                                eway_consolidation_data[seller] = []
                            eway_consolidation_data[seller].append({
                                'dc_data': dc_data_list,
                                'dc_number': result['dc_number'],
                                'vehicle_number': result['vehicle_number'],
                                'trip_refs': result.get('trip_refs', [])
                            })
                            print(f"   🔄 Added to consolidation: {seller} - DC {result['dc_number']} (Vehicle {result['vehicle_number']})")
                            
                        # Track individual e-way template generation success (if not consolidating)
                        if generate_individual_eway:
                            if result.get('eway_template_status') == 'success':
                                eway_success_count += 1
                            elif result.get('eway_template_status') == 'failed':
                                eway_fail_count += 1
                        # Track PDF generation success
                        if generate_pdfs:
                            if result.get('pdf_status') == 'success':
                                pdf_success_count += 1
                            elif result.get('pdf_status') == 'failed':
                                pdf_fail_count += 1
            
            # DEBUG: Show consolidation data summary
            print(f"\n📊 Consolidation Data Summary:")
            for seller, dc_entries in eway_consolidation_data.items():
                vehicle_count = len(set(entry['vehicle_number'] for entry in dc_entries))
                print(f"   {seller}: {len(dc_entries)} DCs from {vehicle_count} vehicles")
                for entry in dc_entries:
                    print(f"     - DC {entry['dc_number']} (Vehicle {entry['vehicle_number']})")
            
            # NEW: Generate consolidated e-way bill templates if requested
            if generate_eway_templates and consolidate_eway_bills and eway_consolidation_data:
                print(f"\n🔄 Generating consolidated e-way bill templates for {len(eway_consolidation_data)} sellers...")
                consolidation_results = self._generate_consolidated_eway_templates(eway_consolidation_data, output_dir)
                
                # Update results with consolidation information
                for seller, consolidation_result in consolidation_results.items():
                    # Find all DCs for this seller and update them with consolidation info
                    seller_results = [r for r in results if r.get('hub_type') == seller]
                    
                    for result in seller_results:
                        if consolidation_result['success']:
                            result.update({
                                'eway_template_generated': True,
                                'eway_template_path': consolidation_result['file_path'],
                                'eway_template_status': 'consolidated_success',
                                'eway_template_type': 'consolidated',
                                'eway_template_seller': seller,
                                'eway_row_count': consolidation_result.get('total_rows', 0)
                            })
                            eway_success_count += 1
                        else:
                            result.update({
                                'eway_template_status': 'consolidated_failed',
                                'eway_template_type': 'consolidated',
                                'eway_template_seller': seller,
                                'eway_error': consolidation_result.get('error', 'Unknown consolidation error')
                            })
                            eway_fail_count += 1
            
            # Enhanced logging
            dc_count = len(results)
            print(f"✅ Generated {dc_count} vehicle-based DCs")
            
            if generate_pdfs:
                print(f"📄 PDF Generation Summary:")
                print(f"   ✅ Successful: {pdf_success_count}")
                if pdf_fail_count > 0:
                    print(f"   ❌ Failed: {pdf_fail_count}")
                    # Show specific failures
                    failed_results = [r for r in results if r.get('pdf_status') == 'failed']
                    for failed in failed_results:
                        print(f"      - {failed['dc_number']}: {failed.get('pdf_error', 'Unknown error')}")
                print(f"   📊 Total PDFs: {pdf_success_count}/{dc_count}")
            
            if generate_eway_templates:
                print(f"📋 E-Way Template Generation Summary:")
                print(f"   ✅ Successful: {eway_success_count}")
                if eway_fail_count > 0:
                    print(f"   ❌ Failed: {eway_fail_count}")
                    # Show specific failures
                    failed_results = [r for r in results if r.get('eway_template_status') in ['failed', 'consolidated_failed']]
                    for failed in failed_results:
                        print(f"      - {failed['dc_number']}: {failed.get('eway_error', 'Unknown error')}")
                
                if consolidate_eway_bills:
                    unique_sellers = len(set(r.get('eway_template_seller') for r in results if r.get('eway_template_type') == 'consolidated'))
                    print(f"   📊 Total E-Way Templates: {unique_sellers} consolidated files (covering {dc_count} DCs)")
                else:
                    print(f"   📊 Total E-Way Templates: {eway_success_count}/{dc_count}")
            
            return results
            
        except Exception as e:
            print(f"❌ Error generating vehicle DCs: {str(e)}")
            return []
    
    def create_generation_summary(self, results, vehicle_assignments):
        """Create a summary of the generation process including PDF and e-way template information"""
        try:
            # Calculate e-way template statistics
            eway_generated_count = sum(1 for r in results if r.get('eway_template_generated', False))
            eway_failed_count = sum(1 for r in results if r.get('eway_template_status') == 'failed')
            eway_success_count = sum(1 for r in results if r.get('eway_template_status') == 'success')
            
            # Calculate PDF statistics
            pdf_generated_count = sum(1 for r in results if r.get('pdf_generated', False))
            pdf_failed_count = sum(1 for r in results if r.get('pdf_status') == 'failed')
            pdf_success_count = sum(1 for r in results if r.get('pdf_status') == 'success')
            
            summary = {
                'generation_timestamp': datetime.now().isoformat(),
                'total_dcs_generated': len(results),
                'vehicles_processed': len(set(r['vehicle_number'] for r in results)),
                'eway_templates_generated': eway_generated_count,
                'eway_templates_successful': eway_success_count,
                'eway_templates_failed': eway_failed_count,
                'pdfs_generated': pdf_generated_count,
                'pdfs_successful': pdf_success_count,
                'pdfs_failed': pdf_failed_count,
                'unified_generation_enabled': eway_generated_count > 0,
                'pdf_generation_enabled': pdf_generated_count > 0,
                'details': []
            }
            
            # Group results by vehicle
            vehicle_groups = {}
            for result in results:
                vehicle_num = result['vehicle_number']
                if vehicle_num not in vehicle_groups:
                    vehicle_groups[vehicle_num] = []
                vehicle_groups[vehicle_num].append(result)
            
            # Create detailed summary
            for vehicle_num, vehicle_results in vehicle_groups.items():
                # Calculate vehicle-specific e-way stats
                vehicle_eway_success = sum(1 for r in vehicle_results if r.get('eway_template_status') == 'success')
                vehicle_eway_failed = sum(1 for r in vehicle_results if r.get('eway_template_status') == 'failed')
                
                # Calculate vehicle-specific PDF stats
                vehicle_pdf_success = sum(1 for r in vehicle_results if r.get('pdf_status') == 'success')
                vehicle_pdf_failed = sum(1 for r in vehicle_results if r.get('pdf_status') == 'failed')
                
                vehicle_detail = {
                    'vehicle_number': vehicle_num,
                    'dcs_generated': len(vehicle_results),
                    'total_products': sum(r['product_count'] for r in vehicle_results),
                    'hub_types': list(set(r['hub_type'] for r in vehicle_results)),
                    'trip_refs': list(set().union(*[r['trip_refs'] for r in vehicle_results])),
                    'eway_templates_successful': vehicle_eway_success,
                    'eway_templates_failed': vehicle_eway_failed,
                    'pdfs_successful': vehicle_pdf_success,
                    'pdfs_failed': vehicle_pdf_failed,
                    'files': []
                }
                
                # Add file information including e-way templates and PDFs
                for r in vehicle_results:
                    file_info = {
                        'dc_number': r['dc_number'],
                        'dc_file_path': r['file_path'],
                        'pdf_generated': r.get('pdf_generated', False),
                        'pdf_path': r.get('pdf_path'),
                        'pdf_status': r.get('pdf_status', 'not_generated'),
                        'eway_template_generated': r.get('eway_template_generated', False),
                        'eway_template_path': r.get('eway_template_path'),
                        'eway_template_status': r.get('eway_template_status', 'not_generated'),
                        'eway_row_count': r.get('eway_row_count', 0)
                    }
                    
                    # Include error information if PDF generation failed
                    if r.get('pdf_status') == 'failed':
                        file_info['pdf_error'] = r.get('pdf_error')
                    
                    # Include error information if e-way generation failed
                    if r.get('eway_template_status') == 'failed':
                        file_info['eway_error'] = r.get('eway_error')
                    
                    vehicle_detail['files'].append(file_info)
                
                summary['details'].append(vehicle_detail)
            
            # Save summary
            summary_file = f"vehicle_generation_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(summary_file, 'w') as f:
                json.dump(summary, f, indent=2)
            
            print(f"✅ Generation summary saved to {summary_file}")
            return summary_file
            
        except Exception as e:
            print(f"❌ Error creating generation summary: {str(e)}")
            return None 