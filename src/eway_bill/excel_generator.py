#!/usr/bin/env python3
"""
Excel Generator for E-Way Bill Templates
Simple utility to generate Excel files from template data
"""

import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def save_data_to_excel(rows, output_file, columns=None):
    """
    Save data rows to Excel file with formatting
    
    Args:
        rows: List of dictionaries containing row data
        output_file: Path to save the Excel file
        columns: Optional list of column names (if None, will use keys from first row)
        
    Returns:
        True if successful, False otherwise
    """
    try:
        # Ensure output directory exists
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
            logger.info(f"Created output directory: {output_dir}")
        
        # Change file extension to xlsx if it's csv
        if output_file.endswith('.csv'):
            output_file = output_file.replace('.csv', '.xlsx')
        
        # If columns not provided, use keys from first row
        if not columns and rows:
            columns = list(rows[0].keys())
        
        # Create a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        
        # Write header row with formatting
        for j, header in enumerate(columns):
            cell = ws.cell(row=1, column=j+1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            
            # Add borders
            thin_border = Border(left=Side(style='thin'), 
                               right=Side(style='thin'), 
                               top=Side(style='thin'), 
                               bottom=Side(style='thin'))
            cell.border = thin_border
        
        # Write data rows
        for i, row_dict in enumerate(rows):
            for j, col in enumerate(columns):
                value = row_dict.get(col, '')
                cell = ws.cell(row=i+2, column=j+1, value=value)
                
                # Add borders to data cells
                cell.border = Border(left=Side(style='thin'), 
                                   right=Side(style='thin'), 
                                   top=Side(style='thin'), 
                                   bottom=Side(style='thin'))
                
                # Set alignment for better readability
                if isinstance(value, (int, float)) or (isinstance(value, str) and value.replace('.', '', 1).isdigit()):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = min(adjusted_width, 40)
        
        # Save the workbook
        wb.save(output_file)
        
        logger.info(f"✅ Saved {len(rows)} rows to Excel file: {output_file}")
        return True
        
    except Exception as e:
        logger.error(f"❌ Error saving Excel file: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def save_eway_bill_to_excel(rows, output_file):
    """
    Save e-way bill template data to Excel with the exact required column format
    
    Args:
        rows: List of dictionaries containing row data
        output_file: Path to save the Excel file
        
    Returns:
        True if successful, False otherwise
    """
    try:
        # Ensure output directory exists
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
            logger.info(f"Created output directory: {output_dir}")
        
        # Change file extension to xlsx if it's csv
        if output_file.endswith('.csv'):
            output_file = output_file.replace('.csv', '.xlsx')
        
        # Define the exact column headers in the required order
        columns = [
            'Supply Type', 'Sub SupplyType', 'Document Type', 'Invoice Reference No',
            'Document Number', 'Document Date', 'Supplier GSTIN', 'Supplier Name',
            'Supplier Address1', 'Supplier Address2', 'Supplier Place', 'Supplier Pincode',
            'Supplier State', 'Customer Billing Name', 'Customer Billing GSTIN',
            'Customer Billing Address 1', 'Customer Billing Address 2', 'Customer Billing City',
            'Customer Billing State', 'Customer Billing Pincode', 'Distance Level (Km)',
            'Transportation Mode (Road/Rail/Air/Ship)', 'Transporter ID', 'Transporter Name',
            'Transportation Date', 'Transporter Doc No', 'Vehicle Number', 'Vehicle Type',
            'Product Name', 'Item Description', 'HSN code', 'Item Quantity',
            'Item Unit of Measurement', 'Taxable Value', 'CGST Rate', 'CGST Amount',
            'SGST Rate', 'SGST Amount', 'IGST Rate', 'IGST Amount', 'CESS Rate',
            'CESS Amount', 'CESS Non Advol Rate', 'CESS Non Advol Tax Amount',
            'Other Value', 'Other Charges - TCS', 'Total Transaction Value',
            'Customer Shipping Name', 'Customer Shipping Address 1',
            'Customer Shipping Address 2', 'Customer Shipping City',
            'Customer Shipping PinCode', 'Customer Shipping State Code',
            'Supplier Dispatch Name', 'Supplier Dispatch Address 1',
            'Supplier Dispatch Address 2', 'Supplier Dispatch Place',
            'Supplier Dispatch PinCode', 'Supplier Dispatch State Code',
            'My Branch', 'Replace E-way bill', 'Is this Supply to/from SEZ unit?',
            'Eway Bill Transaction Type', 'Sub Type Description'
        ]
        
        # Create a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        
        # Add header rows for mandatory/optional/conditional (ClearTax format)
        header_rows = [
            ['', '', 'Mandatory'] + [''] * (len(columns) - 3),
            ['', '', 'Optional'] + [''] * (len(columns) - 3),
            ['', '', 'Conditional'] + [''] * (len(columns) - 3)
        ]
        
        # Write header rows with formatting
        for i, header_row in enumerate(header_rows):
            for j, value in enumerate(header_row):
                cell = ws.cell(row=i+1, column=j+1, value=value)
                cell.font = Font(bold=True)
        
        # Write column names with formatting
        for j, col_name in enumerate(columns):
            cell = ws.cell(row=4, column=j+1, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            
            # Add borders
            thin_border = Border(left=Side(style='thin'), 
                               right=Side(style='thin'), 
                               top=Side(style='thin'), 
                               bottom=Side(style='thin'))
            cell.border = thin_border
        
        # Write data rows
        for i, row_dict in enumerate(rows):
            for j, col in enumerate(columns):
                value = row_dict.get(col, '')
                cell = ws.cell(row=i+5, column=j+1, value=value)
                
                # Add borders to data cells
                cell.border = Border(left=Side(style='thin'), 
                                   right=Side(style='thin'), 
                                   top=Side(style='thin'), 
                                   bottom=Side(style='thin'))
                
                # Set alignment for better readability
                if isinstance(value, (int, float)) or (isinstance(value, str) and value.replace('.', '', 1).isdigit()):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = min(adjusted_width, 40)
        
        # Save the workbook
        wb.save(output_file)
        
        logger.info(f"✅ Saved {len(rows)} rows to e-way bill Excel file: {output_file}")
        return True
        
    except Exception as e:
        logger.error(f"❌ Error saving e-way bill Excel file: {str(e)}")
        import traceback
        traceback.print_exc()
        return False 