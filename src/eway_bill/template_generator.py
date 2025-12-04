import os
import csv
import json
from datetime import datetime

class TemplateGenerator:
    # ClearTax e-way bill template headers
    TEMPLATE_HEADERS = [
        'Supply Type', 'Sub Type', 'Document Type', 'Document No', 'Document Date',
        'From GSTIN', 'From State', 'To GSTIN', 'To State', 'Item Name', 'Description',
        'HSN', 'Unit', 'Quantity', 'Unit Price', 'Taxable Value', 'CGST Rate', 'SGST Rate',
        'IGST Rate', 'CESS Rate', 'CESS Non Advol Rate', 'Transport Mode', 'Distance (in KM)',
        'Transporter Name', 'Transporter ID', 'Transporter Doc No', 'Transporter Doc Date',
        'Vehicle No', 'Vehicle Type'
    ]
    
    def __init__(self, gstin_mapping=None):
        """Initialize the template generator with optional GSTIN mapping."""
        self.gstin_mapping = gstin_mapping or {}
        
    def generate_template_from_dc(self, dc_data):
        """Generate e-way bill template rows from DC data."""
        rows = []
        
        # Extract common data
        document_no = dc_data.get('dc_number', '')
        document_date = dc_data.get('date', '')
        
        # Try to format the date if it's not in the expected format
        try:
            if isinstance(document_date, str) and document_date:
                parsed_date = datetime.strptime(document_date, "%d-%b-%Y")
                document_date = parsed_date.strftime("%d/%m/%Y")
        except ValueError:
            # Keep original format if parsing fails
            pass
            
        from_facility = dc_data.get('from_facility', '')
        to_facility = dc_data.get('to_facility', '')
        
        # Get GSTINs from mapping if available
        from_gstin = self.gstin_mapping.get(from_facility, '')
        to_gstin = self.gstin_mapping.get(to_facility, '')
        
        # Extract state codes from GSTIN (first two digits)
        from_state = from_gstin[:2] if len(from_gstin) >= 2 else ''
        to_state = to_gstin[:2] if len(to_gstin) >= 2 else ''
        
        # Transport details
        vehicle_no = dc_data.get('vehicle_number', '')
        transport_mode = "Road"  # Default to road transport
        distance = dc_data.get('distance', '')  # Changed from '100' to '' for blank column
        
        # Process each product
        for product in dc_data.get('products', []):
            item_name = product.get('item_name', '')
            description = product.get('description', item_name)
            hsn = product.get('hsn', '')
            unit = product.get('unit', 'PCS')
            quantity = product.get('quantity', '0')
            unit_price = product.get('unit_price', '0')
            taxable_value = product.get('taxable_value', '')
            
            # If taxable_value is not provided, calculate it
            if not taxable_value and unit_price and quantity:
                try:
                    taxable_value = str(float(unit_price) * float(quantity))
                except (ValueError, TypeError):
                    taxable_value = '0'
            
            # Tax rates
            cgst_rate = product.get('cgst_rate', '0')
            sgst_rate = product.get('sgst_rate', '0')
            igst_rate = product.get('igst_rate', '0')
            cess_rate = product.get('cess_rate', '0')
            cess_non_advol_rate = product.get('cess_non_advol_rate', '0')
            
            # Create row
            row = [
                'Outward',  # Supply Type
                'Supply',   # Sub Type
                'Delivery Challan',  # Document Type
                document_no,
                document_date,
                from_gstin,
                from_state,
                to_gstin,
                to_state,
                item_name,
                description,
                hsn,
                unit,
                quantity,
                unit_price,
                taxable_value,
                cgst_rate,
                sgst_rate,
                igst_rate,
                cess_rate,
                cess_non_advol_rate,
                transport_mode,
                distance,
                '',  # Transporter Name
                '',  # Transporter ID
                '',  # Transporter Doc No
                '',  # Transporter Doc Date
                vehicle_no,
                'Regular'  # Vehicle Type
            ]
            
            rows.append(row)
            
        return rows
    
    def save_to_csv(self, rows, output_file):
        """Save the generated template rows to a CSV file."""
        # Ensure the output directory exists
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
            
        with open(output_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(self.TEMPLATE_HEADERS)
            writer.writerows(rows)
            
    def load_gstin_mapping(self, mapping_file):
        """Load GSTIN mapping from a JSON file."""
        if os.path.exists(mapping_file):
            with open(mapping_file, 'r') as f:
                self.gstin_mapping = json.load(f)
                return True
        return False
    
    def save_to_excel(self, rows, output_file):
        """Save the generated template rows to an Excel file."""
        # Ensure the output directory exists
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
            
        # Change extension to xlsx if needed
        if output_file.endswith('.csv'):
            output_file = output_file.replace('.csv', '.xlsx')
            
        # Import openpyxl for Excel handling
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Create a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        
        # Write header row with formatting
        for j, header in enumerate(self.TEMPLATE_HEADERS):
            cell = ws.cell(row=1, column=j+1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            
            # Add borders
            thin_border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin'))
            cell.border = thin_border
        
        # Write data rows
        for i, row_data in enumerate(rows):
            for j, value in enumerate(row_data):
                cell = ws.cell(row=i+2, column=j+1, value=value)
                
                # Add borders to data cells
                cell.border = Border(left=Side(style='thin'), 
                                    right=Side(style='thin'), 
                                    top=Side(style='thin'), 
                                    bottom=Side(style='thin'))
                
                # Set alignment for better readability
                if isinstance(value, (int, float)) or (isinstance(value, str) and value.replace('.', '', 1).isdigit()):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = min(adjusted_width, 40)
        
        # Save the workbook
        wb.save(output_file) 