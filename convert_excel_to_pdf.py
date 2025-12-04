#!/usr/bin/env python3
"""
Convert Excel DC to PDF
Reads an existing Excel DC file and converts it to PDF format
"""

import os
import sys
from datetime import datetime
from decimal import Decimal
import pandas as pd
from openpyxl import load_workbook
from src.pdf_generator.dc_pdf_generator import create_dc_pdf_from_excel_data

def read_excel_dc_data(excel_path):
    """
    Read DC data from an existing Excel file
    """
    try:
        print(f"📖 Reading Excel file: {excel_path}")
        
        # Load the workbook
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Extract basic information from the Excel file
        dc_data = {}
        
        # Read header information
        dc_data['serial_number'] = ws['C3'].value or 'Unknown'
        
        # Parse date
        date_str = ws['C4'].value
        if isinstance(date_str, str):
            try:
                dc_data['date'] = datetime.strptime(date_str, '%d-%b-%Y')
            except:
                dc_data['date'] = datetime.now()
        else:
            dc_data['date'] = date_str or datetime.now()
        
        dc_data['vehicle_number'] = ws['I4'].value or 'Unknown'
        
        # Read party information
        dc_data['sender_name'] = ws['B8'].value or 'Unknown Sender'
        dc_data['receiver_name'] = ws['F8'].value or 'Unknown Receiver'
        dc_data['hub_address'] = ws['F9'].value or 'Unknown Address'
        
        # Determine hub type from sender name
        sender_name = dc_data['sender_name'].lower()
        if 'amolakchand' in sender_name:
            dc_data['hub_type'] = 'AMOLAKCHAND'
        elif 'bodega' in sender_name:
            dc_data['hub_type'] = 'BODEGA'
        else:
            dc_data['hub_type'] = 'SOURCINGBEE'
        
        # Read product data starting from row 14
        products = []
        row = 14
        
        while True:
            # Check if we've reached the totals row or end of data
            s_no = ws.cell(row=row, column=1).value
            
            # Stop if we hit the "Total" row or if s_no is None/empty
            if not s_no:
                break
            if isinstance(s_no, str) and s_no.strip().lower() == "total":
                break
                
            product = {
                'Description': ws.cell(row=row, column=2).value or 'Unknown Product',
                'HSN': ws.cell(row=row, column=3).value or 'Unknown',
                'Quantity': Decimal(str(ws.cell(row=row, column=4).value or 0)),
                'Value': Decimal(str(ws.cell(row=row, column=5).value or 0)),
                'GST Rate': Decimal(str(ws.cell(row=row, column=6).value or 0)),
                'Cess': Decimal(str(ws.cell(row=row, column=9).value or 0)) * Decimal('100') / Decimal(str(ws.cell(row=row, column=5).value or 1))  # Convert CESS amount back to rate
            }
            
            products.append(product)
            row += 1
            
            # Safety check to avoid infinite loop - increased limit and added more specific check
            if row > ws.max_row + 10:
                print(f"⚠️ Safety break at row {row} (max_row: {ws.max_row})")
                break
        
        dc_data['products'] = products
        
        print(f"✅ Successfully read DC data:")
        print(f"   Serial Number: {dc_data['serial_number']}")
        print(f"   Date: {dc_data['date'].strftime('%d-%b-%Y')}")
        print(f"   Vehicle: {dc_data['vehicle_number']}")
        print(f"   Hub Type: {dc_data['hub_type']}")
        print(f"   Products: {len(products)}")
        
        return dc_data
        
    except Exception as e:
        print(f"❌ Error reading Excel file: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def main():
    """Convert the specific Excel DC to PDF"""
    
    # The Excel file path
    excel_path = "/Users/jumbotail/Desktop/e-way bill /generated_vehicle_dcs/DC_AKVHDCMYR0001_VehicleKA01JS1234_2,204_22,022_22,044.xlsx"
    
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        return
    
    # Read the Excel data
    dc_data = read_excel_dc_data(excel_path)
    if not dc_data:
        print("❌ Failed to read Excel data")
        return
    
    # Generate PDF filename
    pdf_filename = os.path.basename(excel_path).replace('.xlsx', '.pdf')
    pdf_path = os.path.join('test_output', pdf_filename)
    
    # Ensure output directory exists
    os.makedirs('test_output', exist_ok=True)
    
    # Create PDF
    print(f"\n📄 Creating PDF: {pdf_path}")
    success = create_dc_pdf_from_excel_data(dc_data, pdf_path)
    
    if success:
        print(f"✅ PDF created successfully!")
        print(f"📊 File size: {os.path.getsize(pdf_path):,} bytes")
        print(f"📍 Location: {pdf_path}")
    else:
        print("❌ PDF creation failed")

if __name__ == "__main__":
    main() 